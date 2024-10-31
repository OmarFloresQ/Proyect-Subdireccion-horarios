import json
from openpyxl import load_workbook

# Define the yellow, gray, and red color hex codes
YELLOW_HEX = "FFFFE599"  # Available
GRAY_HEX = "FFCCCCCC"  # Virtual or Not Available
RED_LETTER = "ff000000"  # Red text indicating no availability

# Load the workbook
wb = load_workbook("/home/daniel/PycharmProjects/Horarios/Ejemplo Disponibilidad.xlsx", data_only=True)

# Initialize dictionary to store all teachers' data
info = {}

# Iterate through each sheet
for sheet_name in wb.sheetnames:
    try:
        int(sheet_name)  # Ensuring the sheet name is an integer (representing the teacher's ID)
    except ValueError:
        continue

    sheet = wb[sheet_name]

    # Extract teacher's basic information
    teacher_name = sheet["E10"].value
    teacher_id = sheet["E9"].value
    teacher_email = sheet["G9"].value
    teacher_phone1 = sheet["J9"].value
    teacher_phone2 = sheet["J10"].value

    # Initialize the teacher's data dictionary
    teacher_data = {
        "nombre": teacher_name,
        "email": teacher_email,
        "telefono": [teacher_phone1, teacher_phone2],
        "disponibilidad": {}
    }

    # Define days and time slots
    days = [1,2,3,4,5,6,7]
    hours = [hour + 1 for hour in range(15)]  # C13 is 7-8, C14 is 8-9, etc.

    # Iterate through the schedule table (E13:J27)
    for row in range(13, 28):  # Rows 13 to 27 represent time slots
        time_slot = hours[row - 13]
        teacher_data["disponibilidad"][time_slot] = {}

        for col in range(5, 11):  # Columns E (5) to J (10) represent days (Monday to Saturday)
            day = days[col - 5]
            cell = sheet.cell(row=row, column=col)
            fill_color = cell.fill.start_color.index if cell.fill.start_color else None
            font_color = cell.font.color.rgb if cell.font.color else None

            # Check for yellow cells (available time slots)
            if fill_color == YELLOW_HEX:
                group = cell.value
                if group:
                    teacher_data["disponibilidad"][time_slot][day] = f"{group}"
                else:
                    teacher_data["disponibilidad"][time_slot][day] = "Available"

            # Check for gray cells
            elif fill_color == GRAY_HEX:
                group = cell.value
                if font_color != RED_LETTER and group:
                    teacher_data["disponibilidad"][time_slot][day] = f"Virtual {group}"
                else:
                    teacher_data["disponibilidad"][time_slot][day] = "Virtual Available"

            # For any other color, mark as not available
            else:
                teacher_data["disponibilidad"][time_slot][day] = "Not Available"

    # Store teacher data in the main info dictionary
    info[teacher_id] = teacher_data

# Save the dictionary to a JSON file
with open("teacher_availability.json", "w", encoding="utf-8") as json_file:
    json.dump(info, json_file, ensure_ascii=False, indent=4)
    print(info)

print("Data has been saved to teacher_availability.json.")
