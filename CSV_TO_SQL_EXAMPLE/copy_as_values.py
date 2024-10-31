from openpyxl import load_workbook

# Load the workbook
wb = load_workbook("C:\\Users\\Diego\\Downloads\\Ejemplo Disponibilidad.xlsx", data_only=True)  # Loads with last calculated values
wb_no_formula = load_workbook("C:\\Users\\Diego\\Downloads\\Ejemplo Disponibilidad.xlsx")  # Loads with formulas for reference

# Iterate over each sheet
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    sheet_no_formula = wb_no_formula[sheet_name]

    # Loop through each cell in the sheet
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == "f":  # Check if cell contains a formula
                cell.value = sheet_no_formula[cell.coordinate].value  # Replace formula with its calculated value
                cell.data_type = "n"  # Set data type to numeric or general

# Save the workbook with a new name
wb.save("C:\\Users\\Diego\\Documents\\SIGAF\\DISPONIBILIDAD_VALUES.xlsx")
