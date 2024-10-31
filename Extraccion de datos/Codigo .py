import openpyxl
import pandas as pd
from tabulate import tabulate

"""
Cargar el archivo de Excel"""
excel_dataframe = openpyxl.load_workbook("Copia de Ejemplo Disponibilidad.xlsx") #Tiene que cambiar esta ruta 
# d# e el archivo excel tambien mas abajo utilice otro archivo excel pero puede poner el mismo, ya se lo amrque mas abajo
#y en esta parte tiene que cambiar desde que hoja quiere empezar a sacar los datos en la linea de codigo 74

id_disponibilidad: int = 0

""" 
Es necesario que todos los datos empiecen en el misma rango de columnas y filas 
para que se pueda ejecutar de manera correcta. """

"""
Definir los rangos para filas y columnas"""
fila_inicio_datos = 13
fila_fin_datos = 27
columna_inicio_disponibilidad = 5
columna_fin_disponibilidad = 10

"""
Diccionario de colores para verificar la disponibilidad"""
colores_disponibilidad = {
    "FFFFFFFF": "sin disponibilidad",  # Blanco (sin relleno)
    "FFEA9999": "sin disponibilidad",  # Rojo (#ea9a99)
    "FFCCCCCC": "poca disponibilidad",  # Gris (#cccccc)
    "FFFFE599": "disponible"  # Amarillo (#ffe599)
}

""" 
Diccionario para mapear cada horario a un ids secuenciales"""
horario_ids = {
    "7:00-8:00": 1,
    "8:00-9:00": 2,
    "9:00-10:00": 3,
    "10:00-11:00": 4,
    "11:00-12:00": 5,
    "12:00-13:00": 6,
    "13:00-14:00": 7,
    "14:00-15:00": 8,
    "15:00-16:00": 9,
    "16:00-17:00": 10,
    "17:00-18:00": 11,
    "18:00-19:00": 12,
    "19:00-20:00": 13,
    "20:00-21:00": 14,
    "21:00-22:00": 15
}

"""
Diccionario para los dias de la semana, convertir a ids"""
dias_ids = {
    "LUNES": 1,
    "MARTES": 2,
    "MIÉRCOLES": 3,
    "MIERCOLES": 3,
    "JUEVES": 4,
    "VIERNES": 5,
    "SÁBADO (virtual)": 6,
    "DOMINGO": 7
}

""" 
Inicializar lista para almacenar todos los datos"""
todos_datos = []
Datos = []

""" 
Iterar sobre cada hoja de trabajo, comenzando desde la hoja donde inician los datos en el ejemplo inicia desde la hoja
 1 que es la 0, si se necesita ignorar algunas indicar desde que hoja se quiere empezar """
for nombre_hoja in excel_dataframe.sheetnames[0:]:
    """
    Usar el nombre de la hoja como id_empleado."""
    id_empleado = nombre_hoja

    """
    Seleccionar la hoja actual."""
    hoja = excel_dataframe[nombre_hoja]

    """
    Estraer el periodo corrspondiente"""
    periodos = []
    """
    Periodo con hora comentado"""
    #periodo = hoja.cell(row=7, column= 3).value

    periodo = "2025101" #El periodo seria el mismo 2025/01/01
    periodos.append(periodo)

    """
    Usuario predeterminado"""
    usuarios = []
    usuario = "1"  # El usuario que use es el 1
    usuarios.append(usuario)

    """
    Extraigo los horarios y almacenano en una lista "horarios" recorriendo desde "fila_inicio_datos" hasta 
    "fila_fin_datos" en la columna 3."""
    horarios = []
    for row in range(fila_inicio_datos, fila_fin_datos + 1):
        horario = hoja.cell(row=row, column=3).value
        horarios.append(horario)

    """
    Extraer los días y almacenarlos en una lista "dias" recorriendo desde "columna_inicio_disponibilidad" 
    hasta "columna_fin_disponibilidad" en la fila 12."""
    dias = []
    for col in range(columna_inicio_disponibilidad, columna_fin_disponibilidad + 1):
        dia = hoja.cell(row=12, column=col).value
        dias.append(dia)

    """
    Crear registros para cada combinación de día y horario en la hoja actual."""
    for i, row in enumerate(range(fila_inicio_datos, fila_fin_datos + 1)):
        horario = horarios[i]  # Obtener el horario actual
        id_horario = horario_ids.get(horario, "No definido")  # Obtener el ID del horario

        for j, col in enumerate(range(columna_inicio_disponibilidad, columna_fin_disponibilidad + 1)):
            dia = dias[j]  # Obtener el día correspondiente
            id_dias = dias_ids.get(dia, "No definido")  # Obtener el ID del día

            # Obtener la celda actual
            celda = hoja.cell(row=row, column=col)

            """
            Verificar el color de la celda. Si tiene relleno y un color específico, asignar 
            disponibilidad según el color."""

            if celda.fill and celda.fill.start_color:
                color_celda = celda.fill.start_color.rgb
                disponibilidad = colores_disponibilidad.get(color_celda, "sin disponibilidad")
            else:
                disponibilidad = "sin disponibilidad"  # Sin color
            # Verificar si el texto está en color rojo (#FF0000)
            if celda.font and celda.font.color and celda.font.color.rgb == "FFFF0000":
                disponibilidad = "sin disponibilidad"

            """
            Agregar solo los registros con disponibilidad "disponible" o "poca disponibilidad".
            Asignar ids basado en la disponibilidad."""
            if disponibilidad in ["disponible", "poca disponibilidad"]:
                if disponibilidad == "disponible":
                    id_disponibilidad = 1
                elif disponibilidad == "poca disponibilidad":
                    id_disponibilidad = 2
                todos_datos.append([periodo, id_empleado, id_dias, id_horario, id_disponibilidad, usuario])
                Datos.append([periodo, id_empleado, dia, horario, disponibilidad, usuario])
"""
Crear un DataFrame 'df' con todos los datos de ID y guardarlo como CSV."""
df = pd.DataFrame(todos_datos, columns=["periodo", "id", "dia", "hora", "status", "users_id"])
df.to_csv("disponibilidades_ids.csv", index=False, encoding="utf-8")


# Visualizar datos en una tabla
print(tabulate(df, headers="keys", tablefmt="fancy_grid"))
print("Todos los datos con ids han sido guardados en 'disponibilidades_ids.csv'.")

"""
Descomentar estas líneas para guardar y visualizar los datos completos en un segundo CSV."""
df2 = pd.DataFrame(Datos, columns=["Periodo","ID_Empleado", "Dia", "Horario", "Disponibilidad","Usuario"])
df2.to_csv("disponibilidades.csv", index=False, encoding="utf-8")
print(tabulate(df2, headers="keys", tablefmt="fancy_grid"))
print("Todos los datos completos han sido guardados en 'disponibilidades.csv'.")



#Este es el otro archivo excel que use pero puede usar el mismo, nada mas que tiene que cambiar la hoja
# Cargar el archivo de Excel
excel_dataframe = openpyxl.load_workbook("Ejemplo Disponibilidad.xlsx", data_only=True)  # 'data_only=True' permite obtener el resultado de la fórmula

# Seleccionar la primera hoja (índice 0)
hoja = excel_dataframe[excel_dataframe.sheetnames[0]] #Aqui se indica de cual hoja se sacan los datos, de esta hoja
#Saco los datos de todas las materias, de la tabla grande

"""
Definir los rangos para filas y columnas
"""
fila_inicio_datos = 5
fila_fin_datos = 1188
columna_id_profesor = 2
columna_clave_materia = 5
columna_nombre_materia = 6
columna_nombre_profesor = 3

# Listas para almacenar datos
profesores_ids = []
profesores_name = []
materias = []
claves_materias = []
datos_materias = []

"""
Estraer el periodo corrspondiente"""
periodos = []
periodo = "20250101" #El periodo seria el mismo 2025/01/01
periodos.append(periodo)
"""
Append Usuario"""
usuarios = []
usuario = "1" #El usuario predeterminado
usuarios.append(usuario)

# Recorrer las filas dentro del rango especificado
for row in range(fila_inicio_datos, fila_fin_datos + 1):
    # Extraer datos de cada columna específica y añadirlos a las listas
    profesor_id = hoja.cell(row=row, column=columna_id_profesor).value
    profesores_ids.append(profesor_id)

    profesor_name = hoja.cell(row=row, column=columna_nombre_profesor).value
    profesores_name.append(profesor_name)

    materia_nombre = hoja.cell(row=row, column=columna_nombre_materia).value
    materias.append(materia_nombre)

    clave_materia = hoja.cell(row=row, column=columna_clave_materia).value
    claves_materias.append(clave_materia)

    # Agregar los datos en una lista de registros
    datos_materias.append([periodo, profesor_id, profesor_name, clave_materia, materia_nombre, usuario])

"""
Crear un DataFrame 'df' con todos los datos de ID y guardarlo como CSV.
"""
df = pd.DataFrame(datos_materias, columns=["periodo", "id","Nombre_profesor", "uaprendizaje","Nombre_materia","users_id"])
df.to_csv("materias.csv", index=False, encoding="utf-8")

# Visualizar datos en una tabla
print(tabulate(df, headers="keys", tablefmt="fancy_grid"))
print("Todas las materias han sido guardadas en 'materias.csv'.")
