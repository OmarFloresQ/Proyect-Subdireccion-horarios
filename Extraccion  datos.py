import openpyxl
import pandas as pd

# Cargar el archivo de Excel
excel_dataframe = openpyxl.load_workbook("Copia de Ejemplo Disponibilidad.xlsx")

# Definir los rangos para filas y columnas
fila_inicio_datos = 13
fila_fin_datos = 27
columna_inicio_disponibilidad = 5
columna_fin_disponibilidad = 10

# Diccionario de colores para verificar la disponibilidad
colores_disponibilidad = {
    "FFFFFFFF": "sin disponibilidad",  # Blanco (sin relleno)
    "FFEA9999": "sin disponibilidad",  # Rojo (#ea9a99)
    "FFCCCCCC": "poca disponibilidad",  # Gris (#cccccc)
    "FFFFE599": "disponible"           # Amarillo (#ffe599)
}

# Iterar sobre cada hoja de trabajo, comenzando desde la segunda hoja
for nombre_hoja in excel_dataframe.sheetnames[1:]:
    # Seleccionar la hoja actual
    hoja = excel_dataframe[nombre_hoja]

    # Inicializar lista para almacenar datos de la hoja actual
    data = []

    # Extraer los horarios y días
    horarios = [hoja.cell(row=row, column=3).value for row in range(fila_inicio_datos, fila_fin_datos + 1)]
    dias = [hoja.cell(row=12, column=col).value for col in range(columna_inicio_disponibilidad, columna_fin_disponibilidad + 1)]

    # Crear registros para cada combinación de día y horario en la hoja actual
    for i, row in enumerate(range(fila_inicio_datos, fila_fin_datos + 1)):
        horario = horarios[i]  # Obtener el horario actual

        for j, col in enumerate(range(columna_inicio_disponibilidad, columna_fin_disponibilidad + 1)):
            dia = dias[j]  # Obtener el día correspondiente
            # Obtener la celda actual
            celda = hoja.cell(row=row, column=col)

            # Verificar el color de la celda, si tiene relleno y color específico
            if celda.fill and celda.fill.start_color:
                color_celda = celda.fill.start_color.rgb
                disponibilidad = colores_disponibilidad.get(color_celda, "sin información")
            else:
                # Sin relleno explícito
                disponibilidad = "sin disponibilidad"

            # Agregar el registro al formato deseado
            data.append([dia, horario, disponibilidad])

    # Guardar los datos en un archivo CSV con el nombre de la hoja actual
    df = pd.DataFrame(data, columns=["Día", "Horario", "Disponibilidad"])
    nombre_archivo = f"{nombre_hoja}.csv"
    df.to_csv(nombre_archivo, index=False, encoding="utf-8")

    print(f"Los datos de la hoja '{nombre_hoja}' han sido guardados en '{nombre_archivo}'")

print("Proceso completado para todas las hojas.")
